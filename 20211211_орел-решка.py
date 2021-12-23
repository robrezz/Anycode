#конечный цикл с функцией repeat() орёл/решка
import random #подключаем модуль случайных чисел
import itertools #подключаем модуль итераций
import datetime # подключаем модуль дата-время
import openpyxl # подключаем модуль работы с excel
# создаем и сохраняем книгу на 2 листа---
wb = openpyxl.Workbook()
list=wb.active
list.title = "Sheet 1"
wb.create_sheet(index=1, title="Sheet 2")
#wb.save('20211219 orel-reshka.xlsx')
# ---
c100=itertools.count(1,1) # оформляем количество попыток в цикл
for i in c100:
	random.seed() #рестарт модуля случайных чисел
	now = datetime.datetime.now() # определяем дату-время
	#t=int(input("Сколько раз бросаем монетку?_ ")) #задаём количество попыток
	t=1000
	c2=itertools.repeat('счетчик',times=t) #определяем цикл с повторениями
	orel=0 #счетчик "орла"
	reshka=0 #счетчик "решки"
	z=0 #счетчик "невозможного" (монетка на ребро упала)
	for w in c2: #начало цикла
		r=int(random.random()*100) #определяем диапазон случайной выборки
		if r==50: #проверка условия "невозможного"
			z=z+1
		if r>50: #проверка орёл-решка
			orel=orel+1
		else:
			reshka=reshka+1
		#if orel>reshka: #проверка победителя
			#print(orel,' / ',reshka-z," Выиграл Орёл! ",z," раз/а монетка упала на ребро :)")
			#print(round(orel/t*100,2),'% ',round((reshka-z)/t*100,2),'% ',round(z/t*100,2),'%')
		#else:
			#print(orel,' / ',reshka-z," Выиграла Решка! ",z," раз/а монетка упала на ребро :)")
			#print(round(orel/t*100,2),'% ',round((reshka-z)/t*100,2),'% ',round(z/t*100,2),'%')

		#print('-----------------------------------------------')
		#print (str(now))
	b1 = list.cell(row=i, column=1, value=i)
	b2 = list.cell(row=i, column=2, value=t)
	b3 = list.cell(row=i, column=3, value=orel)
	b4 = list.cell(row=i, column=4, value=reshka-z)
	b5 = list.cell(row=i, column=5, value=z)
	b6 = list.cell(row=i, column=6, value=str(now))
	wb.save('20211219 orel-reshka.xlsx')
	if i>=100:
		break
